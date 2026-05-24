from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Outstanding Invoices"

invoices = [
  {"invoiceNumber":"2300857","poNumber":"PO-30932","invoiceDate":"2025-10-23","candidateNames":"Sami Albudairi","amountPreVat":24300.00,"vat":3645.00,"totalDue":27945.00},
  {"invoiceNumber":"2301003","poNumber":"PO-30940","invoiceDate":"2025-12-18","candidateNames":"Salem Bawazeer; Saleh Baqabl; Abdulaziz Saleh; Ali Baqabl; Osama Khalaf; Mohammed Alhashel; Fahad Albqumi","amountPreVat":97581.85,"vat":14637.28,"totalDue":112219.13},
  {"invoiceNumber":"2301008","poNumber":"PO-31391","invoiceDate":"2025-12-18","candidateNames":"Mohammed Alhabash","amountPreVat":19074.64,"vat":2861.19,"totalDue":21935.83},
  {"invoiceNumber":"2301016","poNumber":"PO-32679","invoiceDate":"2025-12-21","candidateNames":"Ahmed Tapsoba; Ibrahim Alshaikhi; Mohammad Bakkar","amountPreVat":22433.30,"vat":3365.00,"totalDue":25798.30},
  {"invoiceNumber":"2301133","poNumber":"PO-33894","invoiceDate":"2026-01-29","candidateNames":"Mohammed Alkaid; Ali Alzahrani; Razan Aljabri","amountPreVat":26027.83,"vat":3904.17,"totalDue":29932.00},
  {"invoiceNumber":"2301207","poNumber":"PO-33450","invoiceDate":"2026-03-08","candidateNames":"Saleh Aldossary","amountPreVat":13943.48,"vat":2091.52,"totalDue":16035.00},
  {"invoiceNumber":"2301214","poNumber":"PO-33899","invoiceDate":"2026-03-08","candidateNames":"Muhammad Safik","amountPreVat":12453.77,"vat":1868.06,"totalDue":14321.83},
  {"invoiceNumber":"2301226","poNumber":"PO-34611","invoiceDate":"2026-03-08","candidateNames":"Hasihm Wael Alhashmi; Abdulaziz Jumaah; Hassan Alharbi","amountPreVat":41830.43,"vat":6274.57,"totalDue":48105.00},
  {"invoiceNumber":"2301227","poNumber":"PO-34596","invoiceDate":"2026-03-08","candidateNames":"Monthly Salary","amountPreVat":21380.87,"vat":3206.13,"totalDue":24587.00},
  {"invoiceNumber":"2301234","poNumber":"PO-34579","invoiceDate":"2026-03-08","candidateNames":"Latifah Alzahrani; Ahmed Mohammed","amountPreVat":22420.00,"vat":3363.00,"totalDue":25783.00},
  {"invoiceNumber":"2301243","poNumber":"PO-34217","invoiceDate":"2026-03-08","candidateNames":"Riyadh Qumqumji","amountPreVat":3036.59,"vat":455.48,"totalDue":3492.07},
  {"invoiceNumber":"2301249","poNumber":"PO-34979","invoiceDate":"2026-03-12","candidateNames":"Omar Mahbub; Mamduh Al-Dalbahy","amountPreVat":31800.00,"vat":4770.00,"totalDue":36570.00},
  {"invoiceNumber":"2301257","poNumber":"PO-34781","invoiceDate":"2026-03-12","candidateNames":"Monthly Salary","amountPreVat":15280.00,"vat":2292.00,"totalDue":17572.00},
  {"invoiceNumber":"2301283","poNumber":"PO-33448","invoiceDate":"2026-03-28","candidateNames":"Mustafa Ibrahim Abusamra","amountPreVat":95560.00,"vat":14334.00,"totalDue":109894.00},
  {"invoiceNumber":"2301290","poNumber":"PO-34149","invoiceDate":"2026-03-28","candidateNames":"Taif Tahlawi; Abrar Abu Sham","amountPreVat":16732.17,"vat":2509.83,"totalDue":19242.00},
  {"invoiceNumber":"2301291","poNumber":"PO-34032","invoiceDate":"2026-03-28","candidateNames":"Anwar Aljezani","amountPreVat":13014.78,"vat":1952.22,"totalDue":14966.00},
  {"invoiceNumber":"2301301","poNumber":"PO-33444","invoiceDate":"2026-03-28","candidateNames":"Mohammed Al Amri; Mohammed Al Shammari; Mohammed Faisal Obaid","amountPreVat":41613.32,"vat":6242.00,"totalDue":47855.32},
  {"invoiceNumber":"2301359","poNumber":"PO-32100","invoiceDate":"2026-04-19","candidateNames":"Emad Almutairi; Mohammed Zain","amountPreVat":23570.00,"vat":3535.50,"totalDue":27105.50},
  {"invoiceNumber":"2301360","poNumber":"PO-32679","invoiceDate":"2026-04-19","candidateNames":"Ahmed Tapsoba; Ibrahim Alshaikhi","amountPreVat":22215.04,"vat":3331.96,"totalDue":25547.00},
  {"invoiceNumber":"2301361","poNumber":"PO-33669","invoiceDate":"2026-04-19","candidateNames":"Mohammed Duwide; Ahmed Mihi; Mohammed Nasser; Abdulhady Kenaid","amountPreVat":58562.61,"vat":8784.39,"totalDue":67347.00},
  {"invoiceNumber":"2301362","poNumber":"PO-34149","invoiceDate":"2026-04-19","candidateNames":"Taif Tahlawi; Abrar Abu Sham","amountPreVat":16732.17,"vat":2509.83,"totalDue":19242.00},
  {"invoiceNumber":"2301363","poNumber":"PO-34032","invoiceDate":"2026-04-19","candidateNames":"Anwar Aljezani","amountPreVat":13014.78,"vat":1952.22,"totalDue":14966.00},
  {"invoiceNumber":"2301364","poNumber":"PO-34426","invoiceDate":"2026-04-19","candidateNames":"Ahmed Basyoni; Ahmed Alshanqeeti; Ibrahim Alalkami; Khalid Alzahrani; Osman Fadil","amountPreVat":41123.48,"vat":6168.52,"totalDue":47292.00},
  {"invoiceNumber":"2301365","poNumber":"PO-34680","invoiceDate":"2026-04-19","candidateNames":"Alosaimi; Muwafaq Alyahya","amountPreVat":15120.00,"vat":2268.00,"totalDue":17388.00},
  {"invoiceNumber":"2301366","poNumber":"PO-34611","invoiceDate":"2026-04-19","candidateNames":"Hasihm Wael Alhashmi; Abdulaziz Jumaah; Hassan Alharbi","amountPreVat":41830.43,"vat":6274.57,"totalDue":48105.00},
  {"invoiceNumber":"2301367","poNumber":"PO-34596","invoiceDate":"2026-04-19","candidateNames":"Monthly Salary","amountPreVat":21380.87,"vat":3206.13,"totalDue":24587.00},
  {"invoiceNumber":"2301368","poNumber":"PO-34579","invoiceDate":"2026-04-19","candidateNames":"Latifah Alzahrani; Ahmed Mohammed","amountPreVat":22420.00,"vat":3363.00,"totalDue":25783.00},
  {"invoiceNumber":"2301369","poNumber":"PO-33444","invoiceDate":"2026-04-19","candidateNames":"Mohammed Al Amri; Mohammed Al Shammari; Mohammed Faisal Obaid","amountPreVat":41613.32,"vat":6242.00,"totalDue":47855.32},
  {"invoiceNumber":"2301370","poNumber":"PO-34319","invoiceDate":"2026-04-19","candidateNames":"Mohamed Al Habbash","amountPreVat":19151.30,"vat":2872.70,"totalDue":22024.00},
  {"invoiceNumber":"2301371","poNumber":"PO-34317","invoiceDate":"2026-04-19","candidateNames":"Mohamed Al Habbash","amountPreVat":9295.65,"vat":1394.35,"totalDue":10690.00},
  {"invoiceNumber":"2301372","poNumber":"PO-34979","invoiceDate":"2026-04-19","candidateNames":"Omar Mahbub; Mamduh Al-Dalbahy","amountPreVat":31800.00,"vat":4770.00,"totalDue":36570.00},
  {"invoiceNumber":"2301373","poNumber":"PO-34986","invoiceDate":"2026-04-19","candidateNames":"Mawaddah","amountPreVat":10600.00,"vat":1590.00,"totalDue":12190.00},
  {"invoiceNumber":"2301374","poNumber":"PO-34765","invoiceDate":"2026-04-19","candidateNames":"Belal Assaf; Masar Bader; Majed Abdulrahim","amountPreVat":12244.35,"vat":1836.65,"totalDue":14081.00},
  {"invoiceNumber":"2301376","poNumber":"PO-34931","invoiceDate":"2026-04-19","candidateNames":"Obaid Alamri","amountPreVat":9295.65,"vat":1394.35,"totalDue":10690.00},
  {"invoiceNumber":"2301378","poNumber":"PO-34769","invoiceDate":"2026-04-19","candidateNames":"Meshal Alsharef","amountPreVat":9295.65,"vat":1394.35,"totalDue":10690.00},
  {"invoiceNumber":"2301379","poNumber":"PO-34775","invoiceDate":"2026-04-19","candidateNames":"Mohammed Al Thobaiti","amountPreVat":10600.00,"vat":1590.00,"totalDue":12190.00},
  {"invoiceNumber":"2301380","poNumber":"PO-34781","invoiceDate":"2026-04-19","candidateNames":"Monthly Salary","amountPreVat":15280.00,"vat":2292.00,"totalDue":17572.00},
  {"invoiceNumber":"2301381","poNumber":"PO-34993","invoiceDate":"2026-04-19","candidateNames":"Mohammed Faisal Obaid","amountPreVat":15033.04,"vat":2254.96,"totalDue":17288.00},
  {"invoiceNumber":"2301382","poNumber":"PO-35067","invoiceDate":"2026-04-19","candidateNames":"Ramzi Dakheel","amountPreVat":11154.78,"vat":1673.22,"totalDue":12828.00},
  {"invoiceNumber":"2301383","poNumber":"PO-35123","invoiceDate":"2026-04-19","candidateNames":"Mohammed Hussain","amountPreVat":4807.83,"vat":721.17,"totalDue":5529.00},
  {"invoiceNumber":"2301384","poNumber":"PO-35055","invoiceDate":"2026-04-19","candidateNames":"Mohammad Mustafa Nazer","amountPreVat":13943.48,"vat":2091.52,"totalDue":16035.00},
  {"invoiceNumber":"2301385","poNumber":"PO-35056","invoiceDate":"2026-04-19","candidateNames":"Remaining amount","amountPreVat":303.20,"vat":0.00,"totalDue":303.20},
  {"invoiceNumber":"2301408","poNumber":"PO-35463","invoiceDate":"2026-04-23","candidateNames":"Mohammed Ehab; Goudah Badran; Islam Nagi; Bushra Jubara; Salahadin Younis; Abdulrahman Mohammed","amountPreVat":96415.22,"vat":14462.28,"totalDue":110877.50},
  {"invoiceNumber":"2301409","poNumber":"PO-35343","invoiceDate":"2026-04-23","candidateNames":"Badr Mohamed; Reham Shahin; AEDH Almutairi; Osama Khalaf; Khaled Al-Anzi; Mamdouh Al-Harbi; HAMZAH ALAMRI; YOUSEF KHOSHAIM; Saleh Bakarman","amountPreVat":119919.42,"vat":17987.91,"totalDue":137907.33},
  {"invoiceNumber":"2301410","poNumber":"PO-35410","invoiceDate":"2026-04-23","candidateNames":"Manea Alsugoor","amountPreVat":16267.83,"vat":2440.17,"totalDue":18707.50},
  {"invoiceNumber":"2301411","poNumber":"PO-35406","invoiceDate":"2026-04-23","candidateNames":"Amal Alshehri; Wail Edib; MANAR KATEBAH; Mohamad Charaf","amountPreVat":98530.19,"vat":14779.53,"totalDue":113309.72},
  {"invoiceNumber":"2301412","poNumber":"PO-35337","invoiceDate":"2026-04-23","candidateNames":"Ahmed Saied Morsy Badr","amountPreVat":13153.04,"vat":1972.96,"totalDue":15126.00},
  {"invoiceNumber":"2301413","poNumber":"PO-35338","invoiceDate":"2026-04-23","candidateNames":"Abdulhadi Alrashidi","amountPreVat":16996.30,"vat":2549.44,"totalDue":19545.74},
  {"invoiceNumber":"2301426","poNumber":"PO-35067","invoiceDate":"2026-05-05","candidateNames":"Ramzi Dakheel","amountPreVat":93328.35,"vat":13999.25,"totalDue":107327.60},
  {"invoiceNumber":"2301427","poNumber":"PO-35517","invoiceDate":"2026-05-05","candidateNames":"Feras Zuhair; Abdulaziz Alagha","amountPreVat":248839.13,"vat":37325.87,"totalDue":286165.00},
  {"invoiceNumber":"2301428","poNumber":"PO-35492","invoiceDate":"2026-05-05","candidateNames":"JYC CRM Agent","amountPreVat":7420.00,"vat":1113.00,"totalDue":8533.00},
  {"invoiceNumber":"2301440","poNumber":"PO-34866","invoiceDate":"2026-05-13","candidateNames":"Rahaf AlGhamdi","amountPreVat":12084.35,"vat":1812.65,"totalDue":13897.00},
  {"invoiceNumber":"2301441","poNumber":"PO-34866","invoiceDate":"2026-05-13","candidateNames":"Rahaf AlGhamdi","amountPreVat":12084.35,"vat":1812.65,"totalDue":13897.00},
  {"invoiceNumber":"2301442","poNumber":"PO-34866","invoiceDate":"2026-05-13","candidateNames":"Rahaf AlGhamdi","amountPreVat":12084.35,"vat":1812.65,"totalDue":13897.00},
  {"invoiceNumber":"2301443","poNumber":"PO-34866","invoiceDate":"2026-05-13","candidateNames":"Rahaf AlGhamdi","amountPreVat":12084.35,"vat":1812.65,"totalDue":13897.00},
  {"invoiceNumber":"2301444","poNumber":"PO-34866","invoiceDate":"2026-05-13","candidateNames":"Rahaf AlGhamdi","amountPreVat":12084.35,"vat":1812.65,"totalDue":13897.00},
  {"invoiceNumber":"2301445","poNumber":"PO-34866","invoiceDate":"2026-05-13","candidateNames":"Rahaf AlGhamdi","amountPreVat":12084.35,"vat":1812.65,"totalDue":13897.00},
  {"invoiceNumber":"2301446","poNumber":"PO-34866","invoiceDate":"2026-05-13","candidateNames":"Rahaf AlGhamdi","amountPreVat":12084.35,"vat":1812.65,"totalDue":13897.00},
  {"invoiceNumber":"2301447","poNumber":"PO-34866","invoiceDate":"2026-05-13","candidateNames":"Rahaf AlGhamdi","amountPreVat":12084.35,"vat":1812.65,"totalDue":13897.00},
]

MAROON       = "800000"
DARK_MAROON  = "5C1010"
LIGHT_MAROON = "F5EDED"
WHITE        = "FFFFFF"
ALT_GRAY     = "F8F8F8"

def solid_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def arial(size=10, bold=False, italic=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, italic=italic, color=color)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Row 1: Title
ws.merge_cells("A1:G1")
ws.row_dimensions[1].height = 22
c = ws["A1"]
c.value     = "Sela — Outstanding Invoices Import File"
c.font      = arial(size=13, bold=True, color=WHITE)
c.fill      = solid_fill(MAROON)
c.alignment = center

# Row 2: Subtitle
ws.merge_cells("A2:G2")
ws.row_dimensions[2].height = 18
c = ws["A2"]
c.value     = "57 invoices | SOA Outstanding | Generated 2026-05-17"
c.font      = arial(size=10, italic=True, color="5C0000")
c.fill      = solid_fill(LIGHT_MAROON)
c.alignment = center

# Row 3: Headers
headers = ["PO Number", "Invoice Number", "Invoice Date", "Candidate Name",
           "Pre-Vat", "VAT (SAR)", "Total Amount Due"]
ws.row_dimensions[3].height = 18
for col_idx, h in enumerate(headers, start=1):
    c = ws.cell(row=3, column=col_idx, value=h)
    c.font      = arial(size=10, bold=True, color=WHITE)
    c.fill      = solid_fill(DARK_MAROON)
    c.alignment = center

# Data rows
num_fmt = '#,##0.00'
first_data_row = 4

for i, inv in enumerate(invoices):
    row = first_data_row + i
    fill = solid_fill(ALT_GRAY) if i % 2 == 0 else solid_fill(WHITE)
    ws.row_dimensions[row].height = 15

    cells = [
        (inv["poNumber"],      left),
        (inv["invoiceNumber"], left),
        (inv["invoiceDate"],   center),
        (inv["candidateNames"],left),
        (inv["amountPreVat"],  center),
        (inv["vat"],           center),
        (inv["totalDue"],      center),
    ]
    for col_idx, (val, align) in enumerate(cells, start=1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font      = arial(size=10)
        c.fill      = fill
        c.alignment = align
        if col_idx in (5, 6, 7):
            c.number_format = num_fmt

# Totals row
last_data_row = first_data_row + len(invoices) - 1
totals_row    = last_data_row + 1
ws.row_dimensions[totals_row].height = 16

totals_fill = solid_fill(MAROON)
bold_white  = arial(size=10, bold=True, color=WHITE)

c = ws.cell(row=totals_row, column=1, value="TOTAL")
c.font      = bold_white
c.fill      = totals_fill
c.alignment = center

for col_idx in range(2, 8):
    c = ws.cell(row=totals_row, column=col_idx)
    c.fill      = totals_fill
    c.alignment = center
    c.font      = bold_white
    if col_idx in (5, 6, 7):
        col_letter      = get_column_letter(col_idx)
        c.value         = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        c.number_format = num_fmt

# Column widths
col_widths = [18, 18, 16, 55, 18, 16, 20]
for col_idx, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Freeze panes
ws.freeze_panes = "A4"

out_path = "/Users/nessmamamdouh/Desktop/fisheye-v2/sela_outstanding_invoices_import.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
